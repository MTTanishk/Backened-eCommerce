import openpyxl

def getRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_row)

def getColCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_column)

def ReadData(file,sheetname,rnum,cnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.cell(row=rnum,column=cnum).value)

def WriteData(file,sheetname,rnum,cnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row=rnum,column=cnum).value=data
    workbook.save(file)